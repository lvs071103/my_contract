import argparse
import sys
import os
from xlrd import open_workbook
from xlrd.xldate import xldate_as_datetime
import random
from openpyxl import Workbook, load_workbook


# 输出到文件还是终端，输出到文件
def write_file_for_one_col(filename, users, key):
    if not os.path.exists(filename):
        print("生成新的随机用户")
        # 创建一个文件对象
        wb = Workbook()
        ws = wb.active
        ws.title = key
        ws['A1'] = key
        for each in users:
            print(each)
            ws.append([each])
        wb.save(filename)
    else:
        wb = load_workbook(filename)
        # print(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            if key == sheet_name:
                # 如果sheet存在，则删除该sheet，然后更新sheet
                print('{} sheet exists! 更新随机用户'.format(sheet_name))
                wb.remove(wb[sheet_name])
                wb.create_sheet(key)
                ws = wb[key]
                # print(ws)
                ws['A1'] = key
                for each in users:
                    print(each)
                    ws.append([each])
                wb.save(filename)


def write_file_for_muti_col(filename, users, key):
    if not os.path.exists(filename):
        print("生成新的随机用户")
        # 创建一个文件对象
        wb = Workbook()
        ws = wb.active
        ws.title = key
        ws['A1'] = key
        for each in users:
            print(each)
            ws.append([each])
        wb.save(filename)
    else:
        wb = load_workbook(filename)
        # print(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            if key == sheet_name:
                # 如果sheet存在，则删除该sheet，然后更新sheet
                print('{} sheet exists! 更新随机用户'.format(sheet_name))
                wb.remove(wb[sheet_name])
                wb.save(filename)
            else:
                # 如果sheet不同，则新建sheet，并写入数据
                print("创建新的--{}--随机用户数".format(key))
                wb.create_sheet(key)
                ws = wb[key]
                print(ws)
                ws['A1'] = key
                for each in users:
                    print(each)
                    ws.append([each])
                wb.save(filename)


def users_choice(users, num):
    # print(mail_list)
    return random.choices([*users], k=num)


def xls_to_dict_handler(xls_file):
    workbook = open_workbook(xls_file)
    worksheet = workbook.sheet_by_index(0)
    # 列名称数组
    keys = [
        worksheet.cell(0, col_index).value
        for col_index in range(worksheet.ncols)
    ]

    num_rows = worksheet.nrows
    num_cols = worksheet.ncols
    values = []
    # 行处理（0为标题行，从1开始处理）
    for row in range(1, num_rows):
        d = {}
        # 列处理（从0开始处理）
        for col in range(num_cols):
            if keys[col] == 'purchase_date' or \
                    keys[col] == 'scheduled_replacement_date' or \
                    keys[col] == 'discard_date':
                # 如果date字段的值为''则重新赋值为None, 如果值不为''，则将值转换成datetime的时间格式
                if worksheet.cell_value(row, col) == '':
                    datetime_ = None
                else:
                    datetime_ = xldate_as_datetime(
                        worksheet.cell_value(row, col), workbook.datemode)
                d[keys[col]] = datetime_
            else:
                d[keys[col]] = worksheet.cell_value(row, col)
        values.append(d)

    return values


def xls_to_list_handler(xls_file):
    workbook = open_workbook(xls_file)
    worksheet = workbook.sheet_by_index(0)
    # 列名称数组
    keys = [
        worksheet.cell(0, col_index).value
        for col_index in range(worksheet.ncols)
    ]
    num_rows = worksheet.nrows
    num_cols = worksheet.ncols
    random_list = list()
    tmpDct = {}
    for col in range(num_cols):
        # print("列：", col)
        subList = list()
        for row in range(num_rows):
            # print("行：", row)
            if worksheet.cell_value(row, col) == keys[col]:
                continue
            else:
                subList.append(worksheet.cell_value(row, col))
        tmpDct[keys[col]] = subList
    random_list.append(tmpDct)
    random_list.append({'col_len': int(len(keys))})

    return random_list


if __name__ == '__main__':
    # print(len(sys.argv))
    parser = argparse.ArgumentParser(description="随机输出邮件列表")
    parser.add_argument('-i', '--input', required=True, help="指定输入文件")
    parser.add_argument('-o', '--output', required=True, help="指定输出文件")
    parser.add_argument('-n', '--num', required=True, help="指定随机数量")
    args = parser.parse_args()
    if len(sys.argv) <= 6:
        parser.print_help()
    user_list = xls_to_list_handler(args.input)
    write_list = []
    process = []
    muti_process = []
    for item in user_list:
        for k, v in item.items():
            print(type(v))
            if k == 'col_len' and v != 1 and v != 2 and type(v) is list:
                print(11111111)
                process.append(item)
            elif k == 'col_len' and v != 1 and type(v) is list:
                print(222222222)
                muti_process.append(item)
            else:
                print(333333333333333)
                continue

    print("muti: ", muti_process)
    if len(process) == 1:
        for item in process:
            for k, v in item.items():
                choiced_users = users_choice(users=v, num=int(args.num))
                write_file_for_one_col(key=k,
                                       users=choiced_users,
                                       filename=args.output)
    if len(muti_process) > 1:
        for item in muti_process:
            for k, v in item.items():
                choiced_users = users_choice(users=v, num=int(args.num))
                print("sheet_name: ", k, choiced_users)
                write_file_for_muti_col(key=k,
                                        users=choiced_users,
                                        filename=args.output)
